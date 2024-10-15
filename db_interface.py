import openpyxl
import os

def write_to_database(item_name, buy_price, sell_price, quantity):
    # Specify the full path to the Excel file
    file_path = os.path.join("C:\\Users", "SaadH", "OneDrive - TechSolution623", "Desktop", "codeAlpha", "task_2", "transactions.xlsx")
    
    # Load the existing workbook or create a new one
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        # Write header if new file
        ws.append(["ID", "Item Name", "Buy Price", "Sell Price", "Quantity"])

    # Find the next empty row index
    new_row_index = ws.max_row + 1

    # Write new data
    ws.append([new_row_index, item_name, buy_price, sell_price, quantity])

    # Save the workbook
    wb.save(file_path)

def read_from_database():
    # Specify the full path to the Excel file
    file_path = os.path.join("C:\\Users", "SaadH", "OneDrive - TechSolution623", "Desktop", "codeAlpha", "task_2", "transactions.xlsx")
    
    # Check if the file exists
    if not os.path.exists(file_path):
        return []  # Return an empty list if the file doesn't exist

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))

    return all_rows

def remove_from_database(item_name):
    file_path = os.path.join("C:\\Users", "SaadH", "OneDrive - TechSolution623", "Desktop", "codeAlpha", "task_2", "transactions.xlsx")
    
    if not os.path.exists(file_path):
        return False  # File doesn't exist, can't remove

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Find the item in the sheet and remove it
    for row in ws.iter_rows(min_row=2):  # Skip header
        if row[1].value == item_name:  # Check the item name
            ws.delete_rows(row[0].row)  # Delete the entire row
            wb.save(file_path)
            return True  # Item removed successfully

    return False  # Item not found
